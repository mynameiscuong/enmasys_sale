import logging
import calendar
import io
import base64
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from odoo.tools.misc import file_path

_logger = logging.getLogger(__name__)


class CommissionPerformanceReport(models.TransientModel):
    _name = "commission.performance.report"
    _description = "Performance-Based Compensation (Commission) Calculation Sheet for Salespersons and Architects"
    _rec_name = "display_name"

    # filters
    display_name = fields.Char(
        string="Display name",
        default="PHIẾU TÍNH TIỀN LƯƠNG THEO HIỆU QUẢ (HOA HỒNG) NHÂN VIÊN KINH DOANH VÀ KIẾN TRÚC SƯ")
    x_date_from = fields.Date(string="Date from", required=True)
    x_date_to = fields.Date(string="Date to", required=True)
    x_department_ids = fields.Many2many(
        comodel_name="hr.department", string="Departments",
        relation="commission_performance_report_department_rel", column1="report_id", column2="department_id")
    x_sale_employee_ids = fields.Many2many(
        comodel_name="hr.employee", string="Sale-Employees",
        domain=[('active', '=', True), ('user_id', '!=', False)],
        relation="commission_performance_report_sale_employee_rel", column1="report_id", column2="employee_id")
    x_architect_ids = fields.Many2many(
        comodel_name="res.partner", string="Architects",
        domain=[('x_group_id.code', '=', 'KTS'), ('active', '=', True)],
        relation="commission_performance_report_architect_rel", column1="report_id", column2="architect_id")

    # datas
    x_report_data_ids = fields.One2many(
        comodel_name="commission.performance.data.report", string="Report Datas",
        inverse_name="x_report_id")

    # business
    @api.model
    def default_get(self, fields):
        try:
            defaults = super(CommissionPerformanceReport, self).default_get(fields)
            its = datetime.now().date()
            defaults['x_date_from'] = datetime(year=its.year, month=its.month, day=1).date()
            defaults['x_date_to'] = datetime(
                year=its.year, month=its.month, day=calendar.monthrange(month=its.month, year=its.year)[1]).date()
            return defaults
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    @classmethod
    def _generate_report_filters(
            cls, date_from=None, date_to=None, companies=None, departments=None, employees=None, architects=None):
        try:
            report_filters = dict()
            if companies:
                report_filters['by_so_company_ids'] = " AND so.company_id IN (%(company_ids)s)" % {
                    'company_ids': ', '.join(map(str, companies.ids))
                }
            if date_from and date_to:
                report_filters['by_so_date_order'] = """
                     AND so.date_order::date BETWEEN \'%(from_date)s\'::date AND \'%(to_date)s\'::date
                """ % {
                    'from_date': date_from, 'to_date': date_to
                }
            if departments:
                report_filters['by_so_department_ids'] = "WHERE so.x_department_id IN (%(department_ids)s)" % {
                    'department_ids': ', '.join(map(str, departments.ids))
                }
            if architects:
                report_filters['by_architect_ids'] = " AND so.architect_partner_id IN (%(architect_ids)s)" % {
                    'architect_ids': ', '.join(map(str, architects.ids))
                }
            if employees:
                report_filters['by_sale_person_ids'] = " AND so.user_id IN (%(user_ids)s)" % {
                    'user_ids': ', '.join(map(str, employees.user_id.ids))
                }
            return report_filters
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def generate_raw_datas(self):
        try:
            self.ensure_one()
            QUERY_STATEMENT = """
                SELECT	ROW_NUMBER() OVER (ORDER BY so.x_architect_id, so.x_sale_person_id) AS x_order_sequence,
                        so.x_architect_id AS x_architect_id,
                        so.x_sale_person_id AS x_sale_person_id,
                        so.x_sale_resource_name as x_sale_resource_name,
                        so.x_department_id AS x_department_id,
                        SUM(so.x_order_total_amount) AS x_order_total_amount,
                        SUM(so.x_commission_amount) AS x_commission_amount
                FROM ( 
                    SELECT	so.user_id AS x_sale_person_id,
                            NULL AS x_architect_id,
                            he.name AS x_sale_resource_name,
                            he.department_id AS x_department_id,
                            SUM(so.amount_total) AS x_order_total_amount,
                            SUM(so.amount_total * he.x_commission) AS x_commission_amount
                    FROM sale_order AS so
                    LEFT JOIN hr_employee AS he ON so.user_id = he.user_id 
                    WHERE	so.architect_partner_id IS NULL AND so.user_id IS NOT NULL AND he.active IS TRUE
                            %(by_so_company_ids)s
                            %(by_so_date_order)s
                            %(by_sale_person_ids)s
                            %(by_architect_ids)s
                    GROUP BY    so.user_id, he.name, he.department_id, he.x_commission
                    UNION
                    SELECT	NULL AS x_sale_person_id,
                            so.architect_partner_id AS x_architect_id,
                            rp.name AS x_sale_resource_name,
                            he.department_id AS x_department_id,
                            SUM(so.amount_total) AS x_order_total_amount,
                            SUM(so.amount_total * rp.x_commission) AS x_commission_amount
                    FROM sale_order AS so
                    LEFT JOIN res_partner AS rp ON rp.id = so.architect_partner_id
                    LEFT JOIN hr_employee AS he ON rp.id = he.work_contact_id 
                    WHERE	so.architect_partner_id IS NOT NULL  AND rp.active IS TRUE
                            %(by_so_company_ids)s
                            %(by_so_date_order)s
                            %(by_architect_ids)s
                            %(by_sale_person_ids)s
                    GROUP BY    so.architect_partner_id, rp.name, he.department_id, rp.x_commission
                ) AS so
                %(by_so_department_ids)s
                GROUP BY    so.x_architect_id, so.x_sale_person_id, 
                            so.x_sale_resource_name, so.x_department_id
                ORDER BY    so.x_architect_id, so.x_sale_person_id,
                            so.x_sale_resource_name, so.x_department_id
            """
            report_filters = self._generate_report_filters(
                date_from=self.x_date_from, date_to=self.x_date_to,
                companies=self.env.companies, departments=self.x_department_ids,
                employees=self.x_sale_employee_ids, architects=self.x_architect_ids)
            self.env.cr.execute(QUERY_STATEMENT % {
                'by_so_company_ids': report_filters.get('by_so_company_ids', str()),
                'by_so_date_order': report_filters.get('by_so_date_order', str()),
                'by_so_department_ids': report_filters.get('by_so_department_ids', str()),
                'by_architect_ids': report_filters.get('by_architect_ids', str()),
                'by_sale_person_ids': report_filters.get('by_sale_person_ids', str())
            })
            report_datas = self.env.cr.dictfetchall()
            for data in report_datas:
                data.update({
                    'x_report_id': self.id,
                })
            return report_datas
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def generate_report_datas(self):
        try:
            # reset current datas
            self.x_report_data_ids.unlink()
            # try generate report-datas
            raw_datas = self.generate_raw_datas()
            report_datas = self.env['commission.performance.data.report'].create(raw_datas)
            return report_datas
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    def make_exportation(self):
        try:
            # no exportation if no report-datas before
            if not self.x_report_data_ids:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _("No data to make an exportation."),
                        'type': 'warning',
                        'next': {'type': 'ir.actions.act_window_close'},
                    },
                }
            # prepare exported-file values
            report_name_suffix = f"{self.x_date_from.strftime('%d-%m-%Y')} - {self.x_date_to.strftime('%d-%m-%Y')}"
            report_file_name = f"BÁO CÁO HOA HỒNG {report_name_suffix}"
            report_file_template_name = "commission_performance_report_template.xlsx"
            report_file_path = file_path(file_path=f"enmasys_sale/report/{report_file_template_name}")
            workbook = load_workbook(report_file_path)
            worksheet = workbook.active
            # write sub-title
            report_file_sub_title = worksheet.cell(row=2, column=4)
            report_file_sub_title.value = "%(date_from)s - %(date_to)s" % {
                'date_from': self.x_date_from.strftime('%d/%m/%Y'), 'date_to': self.x_date_to.strftime('%d/%m/%Y')
            }
            report_file_sub_title.font = Font(bold=True, size=15, name="Times New Roman")
            # write report-datas content
            self.write_report_datas_content(report_datas=self.x_report_data_ids, worksheet=worksheet)
            # save report file data
            fp = io.BytesIO()
            workbook.save(fp)
            # create new attachment
            new_report_file = self.env['ir.attachment'].sudo().create({
                'name': report_file_name,
                'datas': base64.encodebytes(fp.getvalue()),
                'res_model': self._name,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            # download new created attachment
            return {
                "type": "ir.actions.act_url",
                "name": report_file_name,
                "url": '/web/content/' + str(new_report_file.id) + '?download=true',
                "target": "new",
            }
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)

    @classmethod
    def _make_cell_fit_content(cls, worksheet, cell, width_rate):
        # make cell fit with cell-content by width-rate
        width_rate = width_rate or 1.0
        cell_content_width = len(str(cell.value))
        current_column_width = worksheet.column_dimensions[cell.column_letter].width or 10
        new_column_width = max(current_column_width, cell_content_width) * width_rate
        worksheet.column_dimensions[cell.column_letter].width = new_column_width

    @classmethod
    def write_report_datas_content(cls, report_datas, worksheet, start_row=6, start_column=2):
        try:
            # prepare init values
            if not report_datas:
                return
            fields_collection = {
                start_column: 'x_order_sequence',
                3: 'x_sale_resource_name',
                4: 'x_department_id',
                5: 'x_order_total_amount',
                6: 'x_commission_amount',
                7: 'x_note'
            }
            START_ROW = start_row
            end_column = next(reversed(fields_collection.keys()))
            content_default_font = Font(size=13, name="Times New Roman")
            footer_content_default_font = Font(size=13, name="Times New Roman", bold=True)
            content_default_amount_style = NamedStyle(name='amount_format_style', number_format='#,##0.00')
            default_side = Side(border_style='thin', color="FF000000")
            default_border = Border(left=default_side, right=default_side, top=default_side, bottom=default_side)
            footer_content_default_alignment = Alignment(horizontal='right', vertical='center')
            # writing datas-content
            for data in report_datas:
                for col_position in range(start_column, end_column + 1):
                    data_content_cel = worksheet.cell(row=start_row, column=col_position)
                    # init cell values
                    _horizontal = 'center'
                    _vertical = 'center'
                    _wrapText = False
                    _width_rate = 1.0
                    # write cell-content
                    field_data = fields_collection.get(col_position)
                    if not field_data:
                        continue
                    field_value = data[field_data]
                    if isinstance(data._fields[field_data], fields.Many2one):
                        field_value = field_value['name'] or str()
                        _width_rate = 1.1
                    if isinstance(data._fields[field_data], fields.Float):
                        field_value = field_value or float()
                    else:
                        field_value = field_value or str()
                    data_content_cel.value = field_value
                    # format content-cell
                    if col_position in range(3, 5):
                        _horizontal = 'left'
                    if col_position == end_column:
                        _horizontal = 'justify'
                        _wrapText = True
                    if col_position in range(5, end_column):
                        _horizontal = 'right'
                        # only set content-cell's style content like amount
                        data_content_cel.style = content_default_amount_style
                    data_content_cel.alignment = Alignment(
                        horizontal=_horizontal, wrapText=_wrapText, vertical=_vertical)
                    data_content_cel.font = content_default_font
                    data_content_cel.border = default_border
                    # make cell fit content
                    cls._make_cell_fit_content(worksheet=worksheet, cell=data_content_cel, width_rate=_width_rate)
                # increasing content's row position
                start_row += 1
            # writing datas-footer-content
            for row in worksheet.iter_rows(
                    min_row=start_row, max_row=start_row + 1, min_col=start_column, max_col=end_column):
                for content_cell in row:
                    if content_cell.row == start_row + 1:
                        # only write content if current is end-row
                        footer_content = str()
                        if content_cell.col_idx == 3:
                            footer_content = _("The Total")
                        if content_cell.col_idx == 5:
                            footer_content = sum(report_datas.mapped('x_order_total_amount') or [0])
                            content_cell.style = content_default_amount_style
                        if content_cell.col_idx == 6:
                            footer_content = sum(report_datas.mapped('x_commission_amount') or [0])
                            content_cell.style = content_default_amount_style
                        content_cell.value = footer_content
                    content_cell.font = footer_content_default_font
                    content_cell.alignment = footer_content_default_alignment
                    cls._make_cell_fit_content(worksheet=worksheet, cell=content_cell, width_rate=1.0)
                    # set border for all content-cell
                    content_cell.border = default_border
        except Exception as e:
            _logger.exception(msg=e)
            raise ValidationError(e)
